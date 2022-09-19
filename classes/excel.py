import openpyxl
import xlsxwriter
import os
import json
import sys
import csv
from dotenv import load_dotenv
from datetime import datetime
from classes.intercept_message import InterceptMessage
import classes.globals as g


class Excel(object):
    def __init__(self):
        load_dotenv('.env')
        self.intercept_messages = []
        self.get_config()
        self.load_codes()
        self.get_country_failures()

    def get_config(self):
        # Features - sort results
        try:
            self.sort_results = int(os.getenv('SORT_RESULTS'))
        except Exception as e:
            self.sort_results = 0

        # Features - statuses to include
        try:
            tmp = os.getenv('STATUSES_TO_INCLUDE')
            self.statuses_to_include = tmp.split(",")
        except Exception as e:
            self.statuses_to_include = ["ready"]

        # Get Excel file for input
        self.source_file = os.getenv('SOURCE_FILE')
        self.resource_path = os.path.join(os.getcwd(), "resources")
        self.source_path = os.path.join(self.resource_path, "source")
        self.source_file_path = os.path.join(self.source_path, self.source_file)

        # Get YAML file for output
        self.yaml_file = os.getenv('YAML_FILE')
        self.yaml_path = os.path.join(self.resource_path, "yml")
        self.yaml_file_path = os.path.join(self.yaml_path, self.yaml_file)

        # Get Excel output file file
        self.excel_output_file = os.getenv('EXCEL_OUTPUT')
        now = datetime.now()
        date_string = now.strftime("%Y-%m-%d_%H-%M")
        date_string = now.strftime("%Y-%m-%d")

        self.excel_output_file = self.excel_output_file.replace("{date}", date_string)
        self.excel_path = os.path.join(self.resource_path, "excel")
        self.excel_output_file_path = os.path.join(self.excel_path, self.excel_output_file)

        # Get typos file
        self.typos_file = os.getenv('TYPOS_FILE')
        self.config_path = os.path.join(self.resource_path, "config")
        self.typos_file_path = os.path.join(self.config_path, self.typos_file)

        # Get log folder
        self.typos_file = os.getenv('TYPOS_FILE')
        self.log_path = os.path.join(self.resource_path, "log")
        self.log_file_path = os.path.join(self.log_path, "log.json")

        # For checking of codes exist
        self.codes_file = os.getenv('CODES_FILE')

        # Get sheet name
        self.sheet_name = os.getenv('SHEET_NAME')

    def load_codes(self):
        g.commodities = []
        with open(self.codes_file) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if line_count > 0:
                    pls = row[2]
                    if pls == "80":
                        g.commodities.append(row[1])
                        g.commodities_dict[row[1]] = row[8]
                line_count += 1

        print(f'{line_count} commodity codes have been read.')

    def get_country_failures(self):
        filename = os.path.join(self.config_path, "country_failures.json")
        f = open(filename)
        g.country_failures = json.load(f)

    def read(self):
        print("Reading source Excel file")
        workbook = openpyxl.load_workbook(self.source_file_path)
        sheet = workbook[self.sheet_name]
        row_index = 0
        for row in sheet.iter_rows():
            row_index += 1
            if row_index > 1:
                # print(row_index)
                term = row[0].internal_value
                total_events = row[1].internal_value
                message = row[6].internal_value
                status = row[7].internal_value
                genuine_term = row[8].internal_value

                if "COUNTRY" not in message:
                    term = str(term).strip().lower() if term is not None else ""
                else:
                    term = str(term).strip() if term is not None else ""
                total_events = int(total_events) if total_events is not None else 0
                message = str(message).strip() if message is not None else ""
                status = str(status).strip().lower() if status is not None else ""
                genuine_term = str(genuine_term).strip().lower() if genuine_term is not None else ""

                if status in self.statuses_to_include and message != "":
                    intercept_message = InterceptMessage(term, message, self.typos_file_path)
                    if intercept_message.is_valid:
                        self.intercept_messages.append(intercept_message)
                    if genuine_term != "":
                        terms = genuine_term.split(",")
                        for i in range(0, len(terms)):
                            terms[i] = terms[i].strip()
                        terms = list(set(terms))
                        for term2 in terms:
                            term2 = term2.strip()
                            if term2 != "" and term2 != term:
                                intercept_message = InterceptMessage(term2, message, self.typos_file_path)
                                if intercept_message.is_valid:
                                    self.intercept_messages.append(intercept_message)

        print("Complete")

    def write_yaml(self):
        if self.sort_results == 1:
            self.sort_the_results()
        s = "en:\n"
        for intercept_message in self.intercept_messages:
            if intercept_message.yaml != "":
                s += intercept_message.yaml

        f = open(self.yaml_file_path, "w")
        f.write(s)
        f.close()

    def sort_the_results(self):
        print("Sorting")
        self.intercept_messages = sorted(self.intercept_messages, key=lambda x: x.term, reverse=False)

    def write_excel(self):
        workbook = xlsxwriter.Workbook(self.excel_output_file_path)

        format_bold = workbook.add_format({'bold': True})
        format_bold.set_align('top')
        format_bold.set_align('left')
        format_bold.set_bg_color("#f0f0f0")

        format_wrap = workbook.add_format({'text_wrap': True})
        format_wrap.set_align('top')
        format_wrap.set_align('left')

        sheet = workbook.add_worksheet("Intercept messages")
        widths = [20, 100]
        for i in range(0, len(widths)):
            sheet.set_column(i, i, widths[i])
        sheet.write(0, 0, "Term", format_bold)
        sheet.write(0, 1, "Message", format_bold)

        row_index = 0
        for intercept_message in self.intercept_messages:
            row_index += 1
            if intercept_message.yaml != "":
                sheet.write(row_index, 0, intercept_message.term, format_wrap)
                sheet.write(row_index, 1, intercept_message.message, format_wrap)

        sheet.freeze_panes(1, 0)
        workbook.close()

    def write_erroneous_digits(self):
        my_json = {
            "success_count": len(self.intercept_messages),
            "erroneous_digits": g.erroneous_digits,
            "incorrect_commodities": g.incorrect_commodities,
            "useless_messages": g.useless_messages,
            "typos": g.typos
        }
        out_file = open(self.log_file_path, "w")
        json.dump(my_json, out_file, indent=6)
        out_file.close()
